#!/usr/bin/env python3
"""
ETL Script: Import Data xlsx → Supabase
Sensus Ekonomi 2026 — Kecamatan Tempuran

Setup:
  pip install openpyxl supabase

Cara Pakai:
  1. Isi SUPABASE_URL dan SUPABASE_KEY di bawah
  2. Jalankan: python etl_import.py
  3. Script akan import semua data dari 4 file xlsx ke Supabase
"""

import os
import sys
import json
import openpyxl
from datetime import datetime

# ─── Konfigurasi ─────────────────────────────────────────────────────────────
# Ganti dengan URL dan Service Role Key Supabase Anda
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://YOUR_PROJECT.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "YOUR_SERVICE_ROLE_KEY")

# Path ke file xlsx (relatif dari posisi script ini)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
FILE_REKAP      = os.path.join(BASE_DIR, "rekap_tempuran.xlsx")
FILE_KUALITAS   = os.path.join(BASE_DIR, "kualitas_pendataan.xlsx")
FILE_ANOM_USAHA = os.path.join(BASE_DIR, "anomali_usaha_tempuran.xlsx")
FILE_ANOM_KEL   = os.path.join(BASE_DIR, "anomali_keluarga_tempuran.xlsx")

# ─── Helpers ──────────────────────────────────────────────────────────────────
def safe(val, typ=str):
    if val is None: return None
    try:
        if typ == int:   return int(val) if val != '' else None
        if typ == float: return float(val) if val != '' else None
        return str(val).strip() if str(val).strip() else None
    except (ValueError, TypeError):
        return None

def to_ts(val):
    if val is None: return None
    if isinstance(val, datetime): return val.isoformat()
    try:
        return datetime.strptime(str(val), "%Y-%m-%d %H:%M:%S").isoformat()
    except:
        return None

def read_sheet(filename, sheet_name, header_row=1):
    wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else f'col_{i}' for i, h in enumerate(rows[header_row - 1])]
    data = []
    for row in rows[header_row:]:
        if any(v is not None for v in row):
            data.append(dict(zip(headers, row)))
    wb.close()
    return data

def insert_batch(supabase_client, table, records, batch_size=500):
    total = len(records)
    inserted = 0
    for i in range(0, total, batch_size):
        batch = records[i:i+batch_size]
        result = supabase_client.table(table).insert(batch).execute()
        inserted += len(batch)
        print(f"  ✓ {table}: {inserted}/{total} records")
    return inserted

# ─── Import Functions ─────────────────────────────────────────────────────────

def import_rekap(sb):
    print("\n📋 Importing rekap_tempuran.xlsx...")

    # --- Sheet: alokasi_tugas (info PPL per SLS) ---
    rows = read_sheet(FILE_REKAP, 'alokasi_tugas', header_row=1)
    records = []
    for r in rows:
        if not r.get('idsls'): continue
        records.append({
            'idsls':       safe(r.get('idsls')),
            'email_pml':   safe(r.get('email_pml')),
            'email_ppl':   safe(r.get('email_ppl')),
            'kdprov':      safe(r.get('kdprov')),
            'kdkab':       safe(r.get('kdkab')),
            'kdkec':       safe(r.get('kdkec')),
            'kddesa':      safe(r.get('kddesa')),
            'kdsls':       safe(r.get('kdsls')),
            'kdsubsls':    safe(r.get('kdsubsls')),
            'nmkab':       safe(r.get('nmkab')),
            'nmkec':       safe(r.get('nmkec')),
            'nmdes':       safe(r.get('nmdes')),
            'nmsls':       safe(r.get('nmsls')),
            'nmsubsls':    safe(r.get('nmsubsls')),
            'nm_pml':      safe(r.get('nm_pml')),
            'nm_ppl':      safe(r.get('nm_ppl')),
            'jml_umkm':    safe(r.get('jml_umkm'), int),
            'target_fasih': safe(r.get('target_fasih'), int),
        })
    if records:
        sb.table('alokasi_tugas').delete().neq('id', 0).execute()
        insert_batch(sb, 'alokasi_tugas', records)

    # --- Sheet: progres_capaian ---
    rows = read_sheet(FILE_REKAP, 'progres_capaian', header_row=1)
    records = []
    for r in rows:
        if not r.get('idsls'): continue
        records.append({
            'idsls':          safe(r.get('idsls')),
            'kdprov':         safe(r.get('kdprov')),
            'kdkab':          safe(r.get('kdkab')),
            'kdkec':          safe(r.get('kdkec')),
            'kddesa':         safe(r.get('kddesa')),
            'kdsubsls':       safe(r.get('kdsubsls')),
            'email_ppl':      safe(r.get('email_ppl')),
            'nm_ppl':         safe(r.get('nm_ppl')),
            'nm_pml':         safe(r.get('nm_pml')),
            'email_pml':      safe(r.get('email_pml') or r.get('email_pml.1')),
            'nmsls':          safe(r.get('nmsls')),
            'nmkec':          safe(r.get('nmkec')),
            'nmdesa':         safe(r.get('nmdesa')),
            'relasi_sls':     safe(r.get('relasi_sls')),
            'target_fasih':   safe(r.get('target_fasih'), int),
            'target_simpul':  safe(r.get('target_simpul'), int),
            'progres_didata': safe(r.get('progres_didata'), int),
            'persen_didata':  safe(r.get('persen_didata'), float),
            'draft':          safe(r.get('draft'), int),
            'persen_draft':   safe(r.get('persen_draft'), float),
            'submit':         safe(r.get('submit'), int),
            'persen_submit':  safe(r.get('persen_submit'), float),
        })
    if records:
        sb.table('progres_capaian').delete().neq('id', 0).execute()
        insert_batch(sb, 'progres_capaian', records)

    # --- Sheet: REKAP PPL ---
    wb = openpyxl.load_workbook(FILE_REKAP, read_only=True, data_only=True)
    ws = wb['REKAP PPL']
    ppl_records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        nm_ppl, nm_pml, target, submit = row[0], row[1], row[2], row[3]
        if nm_ppl and nm_ppl != 'Grand Total' and nm_pml:
            ppl_records.append({
                'nm_ppl': safe(nm_ppl),
                'nm_pml': safe(nm_pml),
                'target_simpul': safe(target, int),
                'submit': safe(submit, int),
            })
    wb.close()
    if ppl_records:
        sb.table('rekap_ppl').delete().neq('id', 0).execute()
        insert_batch(sb, 'rekap_ppl', ppl_records)

    # --- Sheet: rekap Data PML ---
    wb = openpyxl.load_workbook(FILE_REKAP, read_only=True, data_only=True)
    ws = wb['rekap Data PML']
    pml_records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        kec, nm_pml, nm_ppl = row[0], row[1], row[2]
        if nm_pml and nm_ppl:
            pml_records.append({'nm_pml': safe(nm_pml), 'nm_ppl': safe(nm_ppl)})
    wb.close()
    if pml_records:
        sb.table('rekap_pml').delete().neq('id', 0).execute()
        insert_batch(sb, 'rekap_pml', pml_records)

def import_kualitas(sb):
    print("\n📊 Importing kualitas_pendataan.xlsx...")

    # --- kualitas_usaha ---
    rows = read_sheet(FILE_KUALITAS, 'kualitas_usaha', header_row=1)
    records = []
    for i, r in enumerate(rows):
        if not r.get('no') and not r.get('kddesa'): continue
        records.append({
            'no':                    safe(r.get('no'), int),
            'kdkab':                 safe(r.get('kdkab')),
            'kdkec':                 safe(r.get('kdkec')),
            'kddesa':                safe(r.get('kddesa')),
            'kdsls':                 safe(r.get('kdsls')),
            'wilayah':               safe(r.get('Wilayah / Kabupaten')),
            'kecamatan':             safe(r.get('Kecamatan')),
            'kelurahan':             safe(r.get('Kelurahan/Desa')),
            'sls_rt':                safe(r.get('SLS / RT')),
            'jml_prelist_usaha':     safe(r.get('Jumlah Prelist Usaha'), int),
            'bku_ditemukan':         safe(r.get('BKU Ditemukan'), int),
            'persen_bku_ditemukan':  safe(r.get('Persen BKU Ditemukan (%)'), float),
            'bku_tutup':             safe(r.get('BKU Tutup'), int),
            'persen_bku_tutup':      safe(r.get('Persen BKU Tutup (%)'), float),
            'bku_ganda':             safe(r.get('BKU Ganda'), int),
            'persen_bku_ganda':      safe(r.get('Persen BKU Ganda (%)'), float),
            'bku_tidak_ditemukan':   safe(r.get('BKU Tidak Ditemukan'), int),
            'persen_bku_hilang':     safe(r.get('Persen BKU Hilang (%)'), float),
            'bku_baru':              safe(r.get('BKU Baru'), int),
            'persen_bku_baru':       safe(r.get('Persen BKU Baru (%)'), float),
            'total_usaha_bku':       safe(r.get('Total Usaha BKU'), int),
            'usaha_keluarga_didata': safe(r.get('Usaha Keluarga Didata'), int),
            'jumlah_usaha_total':    safe(r.get('Jumlah Usaha Total (BKU Ditemukan + BKU Baru + Keluarga)'), int),
            'persen_usaha_total':    safe(r.get('Persen Jumlah Usaha Total (%)'), float),
            'jml_prelist_ub':        safe(r.get('Jumlah Prelist UB'), int),
            'jml_ub_didata':         safe(r.get('Jumlah UB yang Berhasil Didata'), int),
            'persen_ub_didata':      safe(r.get('Persen UB Didata (%)'), float),
            'jml_prelist_umkm':      safe(r.get('Jumlah Prelist UMKM (UM + UMK)'), int),
            'jml_umkm_didata':       safe(r.get('Jumlah UMKM yang Berhasil Didata'), int),
            'persen_umkm_didata':    safe(r.get('Persen UMKM Didata (%)'), float),
            'total_usaha_didata':    safe(r.get('Total Usaha Didata (UB+UM+UMK)'), int),
            'persen_total_skala':    safe(r.get('Persen Total Skala Didata (%)'), float),
        })
    if records:
        sb.table('kualitas_usaha').delete().neq('id', 0).execute()
        insert_batch(sb, 'kualitas_usaha', records)

    # --- kualitas_keluarga (has duplicate columns, use positional) ---
    wb = openpyxl.load_workbook(FILE_KUALITAS, read_only=True, data_only=True)
    ws = wb['kualitas_keluarga']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    records = []
    for row in all_rows[1:]:
        if not row[0]: continue
        records.append({
            'no':                       safe(row[0], int),
            'kdkab':                    safe(row[1]),
            'kdkec':                    safe(row[2]),
            'kddesa':                   safe(row[3]),
            'kdsls':                    safe(row[4]),
            'wilayah':                  safe(row[10]),
            'kecamatan':                safe(row[11]),
            'kelurahan':                safe(row[12]),
            'sls_rt':                   safe(row[13]),
            'prelist_awal':             safe(row[14], int),
            'ditemukan':                safe(row[15], int),
            'persen_ditemukan':         safe(row[16], float),
            'keluarga_baru':            safe(row[17], int),
            'persen_keluarga_baru':     safe(row[18], float),
            'meninggal':                safe(row[19], int),
            'persen_meninggal':         safe(row[20], float),
            'tidak_eligible':           safe(row[21], int),
            'persen_tidak_eligible':    safe(row[22], float),
            'tidak_dapat_ditemui':      safe(row[23], int),
            'persen_tidak_dapat_ditemui': safe(row[24], float),
            'tidak_ditemukan':          safe(row[25], int),
            'persen_tidak_ditemukan':   safe(row[26], float),
            'keluarga_khusus':          safe(row[27], int),
            'persen_keluarga_khusus':   safe(row[28], float),
            'total_hasil':              safe(row[29], int),
            'persen_capaian':           safe(row[30], float),
        })
    if records:
        sb.table('kualitas_keluarga').delete().neq('id', 0).execute()
        insert_batch(sb, 'kualitas_keluarga', records)

    # --- kualitas_art ---
    rows = read_sheet(FILE_KUALITAS, 'kualitas_art', header_row=1)
    records = []
    for r in rows:
        if not r.get('no'): continue
        records.append({
            'no':                     safe(r.get('no'), int),
            'kdkab':                  safe(r.get('kdkab')),
            'kdkec':                  safe(r.get('kdkec')),
            'kddesa':                 safe(r.get('kddesa')),
            'kdsls':                  safe(r.get('kdsls')),
            'wilayah':                safe(r.get('Wilayah / Kabupaten')),
            'kecamatan':              safe(r.get('Kecamatan')),
            'kelurahan':              safe(r.get('Kelurahan/Desa')),
            'sls_rt':                 safe(r.get('SLS / RT')),
            'tinggal_bersama':        safe(r.get('Tinggal Bersama Keluarga'), int),
            'persen_tinggal_bersama': safe(r.get('Persen Tinggal Bersama (%)'), float),
            'art_baru':               safe(r.get('Anggota Keluarga Baru'), int),
            'persen_art_baru':        safe(r.get('Persen ART Baru (%)'), float),
            'meninggal':              safe(r.get('Meninggal'), int),
            'persen_meninggal':       safe(r.get('Persen Meninggal (%)'), float),
            'pindah_dn':              safe(r.get('Pindah Dalam Negeri (DN)'), int),
            'persen_pindah_dn':       safe(r.get('Persen Pindah DN (%)'), float),
            'pindah_ln':              safe(r.get('Pindah Luar Negeri (LN)'), int),
            'persen_pindah_ln':       safe(r.get('Persen Pindah LN (%)'), float),
            'tidak_ditemukan':        safe(r.get('Tidak Ditemukan'), int),
            'persen_tidak_ditemukan': safe(r.get('Persen Tidak Ditemukan (%)'), float),
            'art_khusus':             safe(r.get('Anggota Keluarga Khusus'), int),
            'persen_art_khusus':      safe(r.get('Persen ART Khusus (%)'), float),
            'total_art':              safe(r.get('Total Anggota Keluarga'), int),
            'persen_total_art':       safe(r.get('Persen Total ART (%)'), float),
        })
    if records:
        sb.table('kualitas_art').delete().neq('id', 0).execute()
        insert_batch(sb, 'kualitas_art', records)

def import_anomali(sb):
    print("\n⚠️  Importing anomali_usaha_tempuran.xlsx...")
    wb = openpyxl.load_workbook(FILE_ANOM_USAHA, read_only=True, data_only=True)
    ws = wb['Anomali Usaha']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    for row in all_rows[1:]:
        if not row[0]: continue
        records.append({
            'kdkab':     safe(row[1]),
            'kdkec':     safe(row[2]),
            'kddesa':    safe(row[3]),
            'kdsls':     safe(row[4]),
            'kdsubsls':  safe(row[5]),
            'kabupaten': safe(row[6]),
            'kecamatan': safe(row[7]),
            'kelurahan': safe(row[8]),
            'kode_sls':  safe(row[9]),
            'nm_usaha':  safe(row[10]),
            'nm_ppl':    safe(row[11]),
            'email_ppl': safe(row[12]),
            'nohp_ppl':  safe(row[13]),
            'link_fasih': safe(row[15]),
            'rule_1': safe(row[16]),
            'rule_2': safe(row[17]),
            'rule_3': safe(row[18]),
            'rule_4': safe(row[19]),
            'rule_5': safe(row[20]),
            'rule_6': safe(row[21]),
            'rule_7': safe(row[22]),
            'rule_8': safe(row[23]),
            'tanggal_update': to_ts(row[24]),
        })

    if records:
        sb.table('anomali_usaha').delete().neq('id', 0).execute()
        insert_batch(sb, 'anomali_usaha', records)

    print("\n⚠️  Importing anomali_keluarga_tempuran.xlsx...")
    wb = openpyxl.load_workbook(FILE_ANOM_KEL, read_only=True, data_only=True)
    ws = wb['Anomali Keluarga']
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    records = []
    for row in all_rows[1:]:
        if not row[0]: continue
        records.append({
            'kdkab':     safe(row[1]),
            'kdkec':     safe(row[2]),
            'kddesa':    safe(row[3]),
            'kdsls':     safe(row[4]),
            'kdsubsls':  safe(row[5]),
            'kabupaten': safe(row[6]),
            'kecamatan': safe(row[7]),
            'kelurahan': safe(row[8]),
            'kode_sls':  safe(row[9]),
            'nm_keluarga': safe(row[10]),
            'nm_ppl':    safe(row[11]),
            'email_ppl': safe(row[12]),
            'nohp_ppl':  safe(row[13]),
            'link_fasih': safe(row[15]),
            'r1_cerai':              safe(row[16]),
            'r2_kk_muda':            safe(row[17]),
            'r3_disabilitas':        safe(row[18]),
            'r4_luas_lantai':        safe(row[19]),
            'r5_selisih_pendapatan': safe(row[20]),
            'r6_listrik':            safe(row[21]),
            'r7_anggota_ekstrem':    safe(row[22]),
            'tanggal_update': to_ts(row[23]),
        })

    if records:
        sb.table('anomali_keluarga').delete().neq('id', 0).execute()
        insert_batch(sb, 'anomali_keluarga', records)

# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    try:
        from supabase import create_client
    except ImportError:
        print("❌ supabase tidak terinstall. Jalankan: pip install supabase")
        sys.exit(1)

    if "YOUR_PROJECT" in SUPABASE_URL:
        print("❌ Harap isi SUPABASE_URL dan SUPABASE_KEY di script ini!")
        print("   Atau set environment variable:")
        print("   set SUPABASE_URL=https://xxx.supabase.co")
        print("   set SUPABASE_SERVICE_KEY=your_service_role_key")
        sys.exit(1)

    print("🔗 Connecting to Supabase...")
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    try:
        import_rekap(sb)
        import_kualitas(sb)
        import_anomali(sb)
        print("\n✅ Import selesai! Semua data berhasil diupload ke Supabase.")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback; traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
